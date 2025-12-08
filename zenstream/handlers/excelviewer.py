import streamlit as st
from os.path import sep
from openpyxl import load_workbook, Workbook
from styler import Styler
import pandas as pd


class ExcelViewer:
    @staticmethod
    def show_file(file_name: str):
        st.header("Excel Viewer")
        st.text(file_name.split(sep)[-1])

        workbook: Workbook = load_workbook(file_name)
        active = workbook.active.title

        st.subheader(active)
        df: pd.DataFrame = pd.read_excel(file_name, engine="openpyxl", header=0)
        Styler.show_dataframe("DataFrame", df)

        with st.expander("Metadata"):
            Styler.show_dict(
                "Metadata",
                {
                    "Sheets": str(len(workbook.sheetnames)),
                    "Sheet names": str(workbook.sheetnames),
                    "Active sheet": active,
                },
            )
